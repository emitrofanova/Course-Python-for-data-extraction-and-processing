import json
from openpyxl import load_workbook

#fin = open('data_4.xlsx', 'r', encoding='utf-8')
fout = open('4out.json', 'w', encoding='utf-8')
wb = load_workbook('data_4.xlsx')
sheet = wb["Sheet0"]
dct = {}

for i in range(2,sheet.max_row+1):
    AdmArea = sheet.cell(row = i, column = 5).value
    if AdmArea not in dct:
        dct[AdmArea] = {}
    Distr = sheet.cell(row = i, column = 6).value

    if Distr not in dct[AdmArea]:
        dct[AdmArea][Distr] = []
    Addres = sheet.cell(row = i, column = 7).value
    dct[AdmArea][Distr].append(Addres)

json.dump(dct, fout, ensure_ascii=False)
fout.close()
