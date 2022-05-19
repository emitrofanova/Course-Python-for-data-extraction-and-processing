from openpyxl import load_workbook
fout = open('17out.txt', 'w', encoding='utf-8')
wb = load_workbook('trekking3_6_6_3.xlsx')
sheet1 = wb["Справочник"]
sheet2 = wb["Раскладка"]
dct = {}
for i in range(2,sheet1.max_row+1):
    num = []
    for j in range(2,sheet1.max_column+1):
        val = sheet1.cell(row=i, column=j).value
        if not val:
            num.append(0)
        else:
            num.append(val)
    dct[sheet1.cell(row=i, column=1).value] = num 
result = {}
for i in range(2,sheet2.max_row+1):
    if sheet2.cell(row=i, column=1).value not in result:
        result[sheet2.cell(row=i, column=1).value] = []
        for element in dct[sheet2.cell(row=i, column=2).value]:
            result[sheet2.cell(row=i, column=1).value].append(element*sheet2.cell(row=i, column=3).value/100)
    else:
        for el in range(len(result[sheet2.cell(row=i, column=1).value])):
            result[sheet2.cell(row=i, column=1).value][el] += dct[sheet2.cell(row=i, column=2).value][el] * sheet2.cell(row=i, column=3).value / 100
for el in result:
    print(*map(int,result[el]), file= fout)
fout.close()
