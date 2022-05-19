from openpyxl import load_workbook
from statistics import mean, median
fout = open('14out.txt', 'w', encoding='utf-8')
spisok = []
wb = load_workbook('salaries_6_5_1.xlsx')
sheet = wb["Лист1"]
mediana = []
aver = []
for i in range(2,sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        spisok.append((sheet.cell(row=i, column=j).value))
    mediana.append((median(spisok),sheet.cell(row=i, column=1).value))
    spisok.clear()
for j in range(2,sheet.max_column+1):
    for i in range(2, sheet.max_row+1):
        spisok.append((sheet.cell(row=i, column=j).value))
    aver.append((mean(spisok),sheet.cell(row=1, column=j).value))
    spisok.clear()
print(max(mediana)[1], max(aver)[1], file=fout)
