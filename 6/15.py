from openpyxl import load_workbook

def mykey(tup):
    return tup[1]

def mykey2(tup):
    return tup[0]

fout = open('15out.txt', 'w', encoding='utf-8')
wb = load_workbook('trekking1_6_6_1.xlsx')
sheet = wb["Справочник"]
answ = []
for i in range(2,sheet.max_row+1):
    answ.append((sheet.cell(row=i, column=2).value, sheet.cell(row=i,column=1).value))
answ = sorted(answ, key=mykey)
answ = sorted(answ, key=mykey2, reverse=True)
for i in answ:
    print(i[1], file=fout)
fout.close()
