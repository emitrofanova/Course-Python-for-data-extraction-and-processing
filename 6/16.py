from openpyxl import load_workbook
fout = open('16out.txt', 'w', encoding='utf-8')
wb = load_workbook('trekking2_6_6_2.xlsx')
sheet1 = wb["Справочник"]
sheet2 = wb["Раскладка"]
dct = {}
kkal = 0
b = 0
zh = 0
ug = 0
for i in range(2,sheet1.max_row+1):
    num = []
    for j in range(2,sheet1.max_column+1):
        val = sheet1.cell(row=i, column=j).value
        if not val:
            num.append(0)
        else:
            num.append(val)
    dct[sheet1.cell(row=i, column=1).value] = num 
for i in range(2,sheet2.max_row+1):
    kkal += dct[sheet2.cell(row=i, column=1).value][0] * sheet2.cell(row=i, column=2).value
    b += dct[sheet2.cell(row=i, column=1).value][1] * sheet2.cell(row=i, column=2).value
    zh += dct[sheet2.cell(row=i, column=1).value][2] * sheet2.cell(row=i, column=2).value
    ug += dct[sheet2.cell(row=i, column=1).value][3] * sheet2.cell(row=i, column=2).value
print(int(kkal//100), int(b//100), int(zh//100), int(ug//100), file=fout)
fout.close()
